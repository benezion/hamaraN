#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def is_hebrew_only_with_5_plus_words(text):
    """
    Check if text contains only Hebrew characters, spaces, and basic punctuation,
    AND has 5 or more Hebrew words.
    Returns True if text contains ONLY Hebrew letters (א-ת), spaces, and allowed punctuation,
    and has at least 5 Hebrew words.
    Returns False if text contains digits, English letters, special characters, or has fewer than 5 words.
    """
    if not text or pd.isna(text):
        return False
    
    text_str = str(text).strip()
    
    # Empty string after stripping
    if not text_str:
        return False
    
    # Check each character first
    for char in text_str:
        # Allow Hebrew letters (א-ת)
        if '\u05D0' <= char <= '\u05EA':
            continue
        # Allow basic spaces and very basic punctuation
        if char in ' .':
            continue
        # Reject everything else (digits, English letters, special chars, etc.)
        return False
    
    # Must contain at least one Hebrew letter
    has_hebrew = any('\u05D0' <= char <= '\u05EA' for char in text_str)
    if not has_hebrew:
        return False
    
    # Count Hebrew words (sequences of Hebrew letters)
    # Split by spaces and count words that contain Hebrew letters
    words = text_str.split()
    hebrew_words = []
    
    for word in words:
        # Remove punctuation and check if word contains Hebrew letters
        clean_word = word.replace('.', '').strip()
        if clean_word and any('\u05D0' <= char <= '\u05EA' for char in clean_word):
            hebrew_words.append(clean_word)
    
    # Must have 5 or more Hebrew words
    return len(hebrew_words) >= 5

def create_formatted_excel(df, output_file):
    """Create a professionally formatted Excel file with RTL support"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        if ws is None:
            return False
        ws.title = "Hebrew Only Column D"

        # Set RTL direction
        if hasattr(ws, 'sheet_view') and ws.sheet_view is not None:
            ws.sheet_view.rightToLeft = True

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        cell_font = Font(name='Arial', size=11)
        cell_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border

        # Auto-adjust column widths
        if ws.max_column > 0:
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = ws.cell(row=1, column=col_idx).column_letter
                for row_idx in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows for performance
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        try:
                            if len(str(cell_value)) > max_length:
                                max_length = len(str(cell_value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Add filters
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

        # Save workbook
        wb.save(output_file)
        return True

    except Exception as e:
        print(f"Error creating formatted Excel: {str(e)}")
        return False

def get_words_from_text(text, start_word=0, num_words=5):
    """Extract a sequence of words from Hebrew text starting at start_word position"""
    if not text or pd.isna(text):
        return []
    
    text_str = str(text).strip()
    words = text_str.split()
    
    # Extract Hebrew words only
    hebrew_words = []
    for word in words:
        clean_word = word.replace('.', '').strip()
        if clean_word and any('\u05D0' <= char <= '\u05EA' for char in clean_word):
            hebrew_words.append(clean_word)
    
    # Return the requested slice of words
    end_word = start_word + num_words
    if start_word >= len(hebrew_words):
        return []
    
    return hebrew_words[start_word:end_word]

def count_hebrew_words(text):
    """Count the number of Hebrew words in a text"""
    if not text or pd.isna(text):
        return 0
    
    text_str = str(text).strip()
    words = text_str.split()
    
    hebrew_word_count = 0
    for word in words:
        clean_word = word.replace('.', '').strip()
        if clean_word and any('\u05D0' <= char <= '\u05EA' for char in clean_word):
            hebrew_word_count += 1
    
    return hebrew_word_count

def join_short_lines(df, column_d, min_words=20):
    """Join consecutive lines in Column D if they have fewer than min_words Hebrew words"""
    print(f"\nJoining lines with fewer than {min_words} Hebrew words...")
    
    df_copy = df.copy().reset_index(drop=True)
    rows_to_remove = set()
    joined_count = 0
    
    i = 0
    while i < len(df_copy):
        if i in rows_to_remove:
            i += 1
            continue
        
        current_text = df_copy.iloc[i][column_d]
        current_word_count = count_hebrew_words(current_text)
        
        # If current line has fewer than min_words, try to join with next lines
        if current_word_count < min_words and current_word_count > 0:
            combined_text = str(current_text).strip()
            combined_word_count = current_word_count
            lines_joined = 0
            
            # Look ahead and join consecutive lines until we reach min_words
            j = i + 1
            while j < len(df_copy) and combined_word_count < min_words:
                if j in rows_to_remove:
                    j += 1
                    continue
                
                next_text = df_copy.iloc[j][column_d]
                next_word_count = count_hebrew_words(next_text)
                
                if next_word_count > 0:
                    # Join the texts with a space
                    combined_text += " " + str(next_text).strip()
                    combined_word_count += next_word_count
                    rows_to_remove.add(j)
                    lines_joined += 1
                    
                    print(f"  Joining row {i+1} with row {j+1} (combined words: {combined_word_count})")
                
                j += 1
            
            # Update the current row with the combined text
            if lines_joined > 0:
                df_copy.iloc[i, df_copy.columns.get_loc(column_d)] = combined_text
                joined_count += lines_joined
        
        i += 1
    
    # Remove the joined rows
    df_final = df_copy.drop(index=list(rows_to_remove)).reset_index(drop=True)
    
    print(f"Line joining results:")
    print(f"  Original rows: {len(df_copy):,}")
    print(f"  Lines joined into other lines: {len(rows_to_remove):,}")
    print(f"  Final rows: {len(df_final):,}")
    print(f"  Total join operations: {joined_count}")
    
    return df_final

def remove_sliding_window_duplicates(df, column_d):
    """Remove rows that contain duplicate 5-word sequences in column D"""
    print(f"\nRemoving sliding window duplicates (5-word sequences)...")
    
    # Create a copy to work with
    df_copy = df.copy().reset_index(drop=True)
    rows_to_remove = set()
    seen_sequences = {}  # sequence -> first occurrence row index
    
    total_rows = len(df_copy)
    
    for current_idx in range(total_rows):
        if current_idx in rows_to_remove:
            continue
            
        current_text = df_copy.iloc[current_idx][column_d]
        if not current_text or pd.isna(current_text):
            continue
        
        # Get all Hebrew words from current row
        hebrew_words = []
        words = str(current_text).strip().split()
        for word in words:
            clean_word = word.replace('.', '').strip()
            if clean_word and any('\u05D0' <= char <= '\u05EA' for char in clean_word):
                hebrew_words.append(clean_word)
        
        if len(hebrew_words) < 5:
            continue
        
        # Extract all possible 5-word sequences from current row
        current_sequences = []
        for start_pos in range(len(hebrew_words) - 4):
            sequence = tuple(hebrew_words[start_pos:start_pos + 5])
            current_sequences.append(sequence)
        
        # Check if any of these sequences were seen before
        sequence_found = False
        for sequence in current_sequences:
            if sequence in seen_sequences:
                # This row contains a duplicate sequence, mark for removal
                rows_to_remove.add(current_idx)
                sequence_found = True
                print(f"  Row {current_idx + 1}: Duplicate 5-word sequence found: {' '.join(sequence)}")
                break
        
        # If no duplicate found, record all sequences from this row
        if not sequence_found:
            for sequence in current_sequences:
                if sequence not in seen_sequences:
                    seen_sequences[sequence] = current_idx
    
    # Remove the duplicate rows
    df_deduplicated = df_copy.drop(index=list(rows_to_remove)).reset_index(drop=True)
    
    print(f"Sliding window deduplication results:")
    print(f"  Original rows: {total_rows:,}")
    print(f"  Rows with duplicate 5-word sequences: {len(rows_to_remove):,}")
    print(f"  Remaining rows: {len(df_deduplicated):,}")
    
    return df_deduplicated

def filter_hebrew_only_column_d(input_file, output_file):
    """Filter Excel file to keep only rows where Column D contains only Hebrew words"""
    print(f"Reading Excel file: {input_file}")

    try:
        df = pd.read_excel(input_file)
        print(f"Original file loaded successfully.")
        print(f"Original data shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")

        if len(df.columns) < 4:
            print(f"Error: File must have at least 4 columns (A, B, C, D). Found {len(df.columns)} columns.")
            return None

        # Get Column D (index 3)
        column_d = df.columns[3]
        print(f"Column D name: '{column_d}'")

        # Apply Hebrew-only filter with 5+ words to Column D
        print(f"\nApplying Hebrew-only filter with 5+ words to Column D:")
        print(f"- Keep only rows where Column D contains ONLY Hebrew letters (א-ת)")
        print(f"- Allow spaces and basic punctuation (. only)")
        print(f"- Exclude digits, English letters, special characters")
        print(f"- Must have 5 or more Hebrew words")

        # Filter rows where Column D contains only Hebrew with 5+ words
        hebrew_mask = df[column_d].apply(is_hebrew_only_with_5_plus_words)
        filtered_df = df[hebrew_mask]

        print(f"\nFILTERING RESULTS:")
        print(f"Original rows: {len(df):,}")
        print(f"Rows with Hebrew-only Column D (5+ words): {len(filtered_df):,}")
        print(f"Rows excluded: {len(df) - len(filtered_df):,}")
        print(f"Retention rate: {len(filtered_df)/len(df)*100:.2f}%")

        if len(filtered_df) == 0:
            print("No rows found with Hebrew-only content in Column D!")
            return None

        # 1. Sort by Column D
        print(f"\nSorting data by Column D...")
        filtered_df = filtered_df.sort_values(by=column_d).reset_index(drop=True)
        print(f"Data sorted successfully by Column D")

        # 2. Remove exact duplicates in Column D
        print(f"\nRemoving exact duplicate rows based on Column D...")
        before_exact_dedup = len(filtered_df)
        filtered_df = filtered_df.drop_duplicates(subset=[column_d]).reset_index(drop=True)
        after_exact_dedup = len(filtered_df)
        print(f"Exact duplicate removal results:")
        print(f"  Before: {before_exact_dedup:,} rows")
        print(f"  After: {after_exact_dedup:,} rows")
        print(f"  Removed: {before_exact_dedup - after_exact_dedup:,} exact duplicates")

        # 3. Remove sliding window duplicates (5-word sequences)
        filtered_df = remove_sliding_window_duplicates(filtered_df, column_d)

        # 4. Join lines with fewer than 20 Hebrew words
        before_line_joining = len(filtered_df)
        filtered_df = join_short_lines(filtered_df, column_d, min_words=20)
        after_line_joining = len(filtered_df)

        # Show examples of kept and excluded content
        print(f"\nExamples of KEPT rows (Hebrew-only Column D with 5+ words):")
        hebrew_examples = df[hebrew_mask][column_d].head(10)
        for i, example in enumerate(hebrew_examples, 1):
            print(f"  {i:2d}. '{example}'")

        print(f"\nExamples of EXCLUDED rows (non-Hebrew or <5 words in Column D):")
        non_hebrew_examples = df[~hebrew_mask][column_d].head(10)
        for i, example in enumerate(non_hebrew_examples, 1):
            if pd.notna(example):
                print(f"  {i:2d}. '{example}' ← Contains non-Hebrew characters or <5 words")

        # Show sample of final filtered and processed data
        print(f"\nSample of final processed data (first 5 rows):")
        for idx, (_, row) in enumerate(filtered_df.head(5).iterrows()):
            print(f"Row {idx+1}:")
            for col in df.columns:
                value = str(row[col])
                if col == column_d:
                    print(f"  {col}: '{value}' ← Hebrew-only (5+ words) ✓ SORTED")
                else:
                    print(f"  {col}: '{value}'")
            print("-" * 50)

        # Create formatted Excel file
        success = create_formatted_excel(filtered_df, output_file)

        if success:
            print(f"\nFormatted Excel file created successfully: {output_file}")
        else:
            # Fallback to basic Excel
            filtered_df.to_excel(output_file, index=False)
            print(f"Basic Excel file created: {output_file}")

        # Create summary text file
        summary_file = output_file.replace('.xlsx', '_summary.txt')
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write("HEBREW-ONLY COLUMN D FILTER SUMMARY\n")
            f.write("="*50 + "\n")
            f.write(f"Processing Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Input file: {input_file}\n")
            f.write(f"Output file: {output_file}\n")
            f.write("-" * 50 + "\n")
            f.write(f"Original rows: {len(df):,}\n")
            f.write(f"After Hebrew filtering: {before_exact_dedup:,}\n")
            f.write(f"After exact duplicate removal: {after_exact_dedup:,}\n")
            f.write(f"After sliding window dedup: {before_line_joining:,}\n")
            f.write(f"Final rows (after line joining): {len(filtered_df):,}\n")
            f.write(f"Total excluded rows: {len(df) - len(filtered_df):,}\n")
            f.write(f"Final retention rate: {len(filtered_df)/len(df)*100:.2f}%\n")
            f.write("="*50 + "\n")
            f.write("\nProcessing Steps Applied:\n")
            f.write("1. Hebrew-only filtering:\n")
            f.write("   - Column D must contain ONLY Hebrew letters (א-ת)\n")
            f.write("   - Spaces and basic punctuation (.) allowed\n")
            f.write("   - Must have 5 or more Hebrew words\n")
            f.write("   - NO digits, English letters, or special characters\n")
            f.write("   - Empty cells excluded\n")
            f.write("2. Sorting by Column D (alphabetical)\n")
            f.write("3. Exact duplicate removal (Column D)\n")
            f.write("4. Sliding window duplicate removal (5-word sequences)\n")
            f.write("5. Line joining (combine lines with <20 Hebrew words)\n")

        print(f"Summary saved to: {summary_file}")

        # Print final statistics
        print("\n" + "="*60)
        print("HEBREW-ONLY COLUMN D FILTER STATISTICS")
        print("="*60)
        print(f"Input file: {input_file}")
        print(f"Output file: {output_file}")
        print(f"Processing Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Original rows: {len(df):,}")
        print(f"After Hebrew filtering: {before_exact_dedup:,}")
        print(f"After exact duplicate removal: {after_exact_dedup:,}")
        print(f"After sliding window dedup: {before_line_joining:,}")
        print(f"Final rows (after line joining): {len(filtered_df):,}")
        print(f"Total excluded rows: {len(df) - len(filtered_df):,}")
        print(f"Final retention rate: {len(filtered_df)/len(df)*100:.2f}%")
        print("="*60)

        return {
            'original_rows': len(df),
            'after_hebrew_filter': before_exact_dedup,
            'after_exact_dedup': after_exact_dedup,
            'after_sliding_window': before_line_joining,
            'final_rows': len(filtered_df),
            'total_excluded': len(df) - len(filtered_df),
            'final_retention_rate': len(filtered_df)/len(df)*100,
            'summary_file': summary_file
        }

    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        return None
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    input_file = "all_2_words_fixed.xlsx"
    output_file = "hebrew_only_column_d.xlsx"

    print("HEBREW-ONLY COLUMN D FILTER (5+ WORDS) + ADVANCED DEDUPLICATION + LINE JOINING")
    print("Filters Excel file to keep only rows where Column D contains ONLY Hebrew words with 5+ words")
    print("Includes sorting, exact duplicate removal, sliding window deduplication, and line joining")
    print("\nFiltering Criteria:")
    print("✓ Column D must contain ONLY Hebrew letters (א-ת)")
    print("✓ Spaces and basic punctuation (.) allowed")
    print("✓ Must have 5 or more Hebrew words")
    print("✗ NO digits (0-9)")
    print("✗ NO English letters (a-z, A-Z)")
    print("✗ NO special characters (!@#$%^&*()_+-={}[]|\\:;\"'<>?,/)")
    print("✗ NO empty cells")
    print("✗ NO rows with fewer than 5 Hebrew words")
    print("\nProcessing Steps:")
    print("1. Hebrew-only filtering")
    print("2. Sort by Column D (alphabetical)")
    print("3. Remove exact duplicates in Column D")
    print("4. Remove sliding window duplicates (5-word sequences)")
    print("5. Join lines with fewer than 20 Hebrew words")
    print("\nFeatures: RTL direction, professional formatting, column filters")
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print("-" * 60)

    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' does not exist in the current directory.")
        print("Available Excel files in current directory:")
        excel_files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xls'))]
        if excel_files:
            for i, file in enumerate(excel_files, 1):
                print(f"  {i}. {file}")
            print(f"\nYou can modify the input_file variable in the script to use one of these files.")
        else:
            print("  No Excel files found in current directory.")
        return

    stats = filter_hebrew_only_column_d(input_file, output_file)

    if stats:
        print("\nProcessing and formatting features applied:")
        print("✓ Hebrew-only content filtering (Column D)")
        print("✓ Column D sorted alphabetically")
        print("✓ Exact duplicate removal")
        print("✓ Sliding window deduplication (5-word sequences)")
        print("✓ Line joining (combine lines with <20 Hebrew words)")
        print("✓ Right-to-left (RTL) direction")
        print("✓ Professional header styling")
        print("✓ Auto-adjusted column widths")
        print("✓ Filter buttons on all columns")
        print("✓ Cell borders and alignment")
        print("✓ All columns preserved")
        print("✓ Summary statistics file created")

if __name__ == "__main__":
    main()
