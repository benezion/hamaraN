#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys
import subprocess
import os
import json

def load_backup_params():
    """Load parameters from backup file if it exists"""
    backup_file = 'extract_params.bu'
    if os.path.exists(backup_file):
        try:
            with open(backup_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_backup_params(start_line, total_lines, from_line, show_lines):
    """Save parameters to backup file"""
    backup_file = 'extract_params.bu'
    params = {
        'start_line': start_line,
        'total_lines': total_lines,
        'from_line': from_line,
        'show_lines': show_lines
    }
    try:
        with open(backup_file, 'w', encoding='utf-8') as f:
            json.dump(params, f, indent=2)
    except Exception as e:
        print(f"Warning: Could not save backup parameters: {str(e)}")

def generate_html_output(extracted_lines, actual_start_line, actual_end_line, output_file):
    """Generate HTML output with line numbers"""
    html_file = output_file.replace('.txt', '.html')
    
    # Create HTML content with proper formatting
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Extracted Lines</title>
    <style>
        body {{
            font-family: 'Courier New', monospace;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .header {{
            background-color: #2c3e50;
            color: white;
            padding: 10px;
            border-radius: 4px;
            margin-bottom: 20px;
        }}
        .line {{
            display: flex;
            border-bottom: 1px solid #eee;
            padding: 2px 0;
        }}
        .line-number {{
            background-color: #ecf0f1;
            color: #2c3e50;
            padding: 2px 8px;
            margin-right: 10px;
            min-width: 60px;
            text-align: right;
            border-radius: 3px;
            font-weight: bold;
        }}
        .line-content {{
            flex: 1;
            padding: 2px 8px;
            white-space: pre-wrap;
            word-wrap: break-word;
        }}
        .active {{
            background-color: #e8f5e8;
        }}
        .empty {{
            color: #bdc3c7;
            font-style: italic;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>Extracted Lines</h2>
            <p>Lines {actual_start_line} to {actual_end_line} | Total: {len(extracted_lines)}</p>
        </div>
        <div class="content">
"""
    
    # Add lines with proper numbering
    for i, line in enumerate(extracted_lines):
        actual_line_number = actual_start_line + i
        
        line_class = "active"
        display_content = line if line.strip() else "[Empty Line]"
        
        if not line.strip():
            content_class = "empty"
        else:
            content_class = ""
        
        # Escape HTML characters in the content
        display_content = display_content.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            
        html_content += f"""            <div class="line {line_class}">
                <div class="line-number">{actual_line_number}</div>
                <div class="line-content {content_class}">{display_content}</div>
            </div>
"""
    
    html_content += """        </div>
    </div>
</body>
</html>"""
    
    # Write HTML file
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    return html_file

def parse_parameters(args, backup_params):
    """Parse command line parameters, supporting comma-separated format with empty values"""
    if len(args) < 2:
        return None, None, None, None
    
    param_str = args[1]
    
    # Check if it's comma-separated
    if ',' in param_str:
        parts = param_str.split(',')
        if len(parts) > 4:
            raise ValueError("Too many comma-separated parameters")
        
        # Pad with empty strings if needed
        while len(parts) < 4:
            parts.append('')
        
        # Parse each part, using backup if empty
        start_line = int(parts[0]) if parts[0].strip() else backup_params.get('start_line')
        total_lines = int(parts[1]) if parts[1].strip() else backup_params.get('total_lines')
        from_line = int(parts[2]) if parts[2].strip() else backup_params.get('from_line', 1)
        show_lines = int(parts[3]) if parts[3].strip() else backup_params.get('show_lines', total_lines if total_lines else 1)
        
        return start_line, total_lines, from_line, show_lines
    
    # Handle space-separated parameters (existing behavior)
    start_line = int(args[1]) if len(args) > 1 else backup_params.get('start_line')
    total_lines = int(args[2]) if len(args) > 2 else backup_params.get('total_lines')
    from_line = int(args[3]) if len(args) > 3 else backup_params.get('from_line', 1)
    show_lines = int(args[4]) if len(args) > 4 else backup_params.get('show_lines', total_lines if total_lines else 1)
    
    return start_line, total_lines, from_line, show_lines

def main():
    try:
        # Load backup parameters
        backup_params = load_backup_params()
        
        # Check command line arguments
        if len(sys.argv) < 2:
            if backup_params:
                print("Using backup parameters:")
                print(f"  start_line: {backup_params.get('start_line', 'N/A')}")
                print(f"  total_lines: {backup_params.get('total_lines', 'N/A')}")
                print(f"  from_line: {backup_params.get('from_line', 'N/A')}")
                print(f"  show_lines: {backup_params.get('show_lines', 'N/A')}")
                start_line = backup_params.get('start_line')
                total_lines = backup_params.get('total_lines')
                from_line = backup_params.get('from_line')
                show_lines = backup_params.get('show_lines')
                if start_line is None or total_lines is None or from_line is None or show_lines is None:
                    print("Error: Incomplete backup parameters")
                    print("Usage: python extract_lines.py <start_line> <total_lines> <from_line> <show_lines>")
                    print("Comma format: python extract_lines.py <start_line>,<total_lines>,<from_line>,<show_lines>")
                    print("Special: python extract_lines.py + | - | =")
                    sys.exit(1)
            else:
                print("Usage: python extract_lines.py <start_line> <total_lines> <from_line> <show_lines>")
                print("Example: python extract_lines.py 680 20 3 5")
                print("  start_line: Starting line number in input file")
                print("  total_lines: Number of lines to extract from input (buffer size)")
                print("  from_line: Starting line within the buffer to display (1-based)")
                print("  show_lines: Number of lines to display from the buffer")
                print("  Result: Extract 20 lines from line 680, then show 5 lines starting from line 3 in buffer")
                print("")
                print("Comma format examples:")
                print("  680,20,3,5  : Extract 20 lines from 680, show 5 lines starting from buffer line 3")
                print("  ,,1,        : Use backup start_line and total_lines, from_line=1, use backup show_lines")
                print("  680,,,5     : start_line=680, use backup total_lines and from_line, show_lines=5")
                print("")
                print("Special commands:")
                print("  +: Move forward by last total_lines, use last total_lines, from_line=1")
                print("  -: Move backward by last total_lines, use last total_lines, from_line=1")
                print("  =: Run the last command as is with the last parameters that saved")
                sys.exit(1)
        else:
            # Check for special commands
            if sys.argv[1] in ['+', '-', '=']:
                if not backup_params or backup_params.get('start_line') is None or backup_params.get('total_lines') is None:
                    print("Error: No backup parameters available for special commands")
                    print("Run the script with normal parameters first")
                    sys.exit(1)
                
                last_start_line = backup_params.get('start_line')
                last_total_lines = backup_params.get('total_lines')
                last_from_line = backup_params.get('from_line', 1)
                last_show_lines = backup_params.get('show_lines', last_total_lines)
                
                if sys.argv[1] == '+':
                    # Add the previous total_lines to previous start_line
                    start_line = last_start_line + last_total_lines
                    total_lines = last_total_lines  # Keep previous total_lines 
                    from_line = 1  # Set current from_line to 1
                    show_lines = last_total_lines  # Set current show_lines to previous total_lines
                    print(f"+ command: Moving forward by {last_total_lines} lines")
                    
                elif sys.argv[1] == '-':
                    # Subtract the previous total_lines from previous start_line
                    start_line = last_start_line - last_total_lines
                    total_lines = last_total_lines  # Keep previous total_lines
                    from_line = 1  # Set current from_line to 1
                    show_lines = last_total_lines  # Set current show_lines to previous total_lines
                    print(f"- command: Moving backward by {last_total_lines} lines")
                    
                elif sys.argv[1] == '=':
                    # Run the last command as is with the last parameters
                    start_line = last_start_line
                    total_lines = last_total_lines
                    from_line = last_from_line
                    show_lines = last_show_lines
                    print(f"= command: Repeating last run with same parameters")
                
                print(f"Special command result: start_line={start_line}, total_lines={total_lines}, from_line={from_line}, show_lines={show_lines}")
                
            else:
                # Parse parameters (supports both comma and space format)
                try:
                    start_line, total_lines, from_line, show_lines = parse_parameters(sys.argv, backup_params)
                    
                    # Fill missing parameters from backup
                    if start_line is None:
                        start_line = backup_params.get('start_line')
                    if total_lines is None:
                        total_lines = backup_params.get('total_lines')
                    if from_line is None:
                        from_line = backup_params.get('from_line', 1)
                    if show_lines is None:
                        show_lines = backup_params.get('show_lines', total_lines if total_lines else 1)
                    
                    if start_line is None or total_lines is None:
                        print("Error: Missing required parameters and no backup data available")
                        print("Usage: python extract_lines.py <start_line> <total_lines> <from_line> <show_lines>")
                        print("Comma format: python extract_lines.py <start_line>,<total_lines>,<from_line>,<show_lines>")
                        sys.exit(1)
                    
                    # Show what parameters were used
                    if ',' in sys.argv[1]:
                        print(f"Comma format parsed: start_line={start_line}, total_lines={total_lines}, from_line={from_line}, show_lines={show_lines}")
                        
                except ValueError as e:
                    print(f"Error: {str(e)}")
                    print("Parameters must be integers")
                    sys.exit(1)
        
        if start_line < 1:
            print("Error: Start line must be 1 or greater")
            sys.exit(1)
        
        if total_lines < 1:
            print("Error: Total lines must be 1 or greater")
            sys.exit(1)
            
        if from_line < 1:
            print("Error: From line must be 1 or greater")
            sys.exit(1)
        
        if show_lines < 1:
            print("Error: Show lines must be 1 or greater")
            sys.exit(1)
        
        if from_line > total_lines:
            print(f"Error: From line ({from_line}) cannot be greater than total lines ({total_lines})")
            sys.exit(1)
        
        # This auto-adjustment logic is not needed since show_lines is properly handled in parse_parameters
        # Removed to prevent overriding explicitly set show_lines values
        
        if from_line + show_lines - 1 > total_lines:
            print(f"Error: From line ({from_line}) + show lines ({show_lines}) - 1 cannot exceed total lines ({total_lines})")
            sys.exit(1)
        
        # Calculate actual extraction range
        actual_start_line = start_line  # Start from start_line in input file
        actual_end_line = actual_start_line + total_lines - 1  # Extract total_lines from actual_start_line
        lines_to_extract = total_lines  # Always extract exactly total_lines
        
        # Calculate display range within the extracted buffer
        display_start_index = from_line - 1  # Convert to 0-based index
        display_end_index = display_start_index + show_lines - 1  # End index for display
        
        # Save current parameters to backup file
        save_backup_params(start_line, total_lines, from_line, show_lines)
        
        # Input and output file names
        input_file = 'all_2_words_filtered_final_complete_all.xlsx'
        output_file = 'sample_input.txt'
        
        # Check if input file exists
        if not os.path.exists(input_file):
            print(f"Error: Input file '{input_file}' not found")
            sys.exit(1)
        
        print(f"Reading from: {input_file}")
        print(f"Parameters: start_line={start_line}, total_lines={total_lines}, from_line={from_line}, show_lines={show_lines}")
        print(f"Extracting lines {actual_start_line} to {actual_end_line} ({lines_to_extract} lines) from input file")
        print(f"Displaying lines {from_line} to {from_line + show_lines - 1} from the extracted buffer")
        print(f"Output file: {output_file}")
        
        # Read the Excel file
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        # Get the data from column D (Context) - extract full buffer first
        extracted_buffer = []
        lines_extracted = 0
        
        # Note: Excel rows are 1-indexed, openpyxl rows are also 1-indexed
        # If the actual_start_line refers to data rows (excluding header), we add 1
        # Assuming row 1 is header, so data starts from row 2
        excel_start_row = actual_start_line + 1  # +1 to skip header
        
        for row_num in range(excel_start_row, excel_start_row + lines_to_extract):
            cell_value = ws.cell(row=row_num, column=4).value  # Column D is column 4
            
            if cell_value is not None:
                extracted_buffer.append(str(cell_value))
                lines_extracted += 1
            else:
                # If cell is empty, add empty line
                extracted_buffer.append("")
                lines_extracted += 1
            
            # Stop if we've reached the end of the worksheet
            if row_num > ws.max_row:
                break
        
        wb.close()
        
        # Now extract the lines to display from the buffer
        lines_to_display = extracted_buffer[display_start_index:display_start_index + show_lines]
        
        # Write to output file (only the lines to display)
        with open(output_file, 'w', encoding='utf-8') as f:
            for line in lines_to_display:
                f.write(line + '\n')
        
        print(f"Successfully extracted {lines_extracted} lines from input file")
        print(f"Displaying {len(lines_to_display)} lines to {output_file}")
        
        # Calculate the actual line numbers for the displayed lines
        display_actual_start = actual_start_line + display_start_index
        display_actual_end = display_actual_start + len(lines_to_display) - 1
        
        # Generate HTML output with line numbers (only for displayed lines)
        html_file = generate_html_output(lines_to_display, display_actual_start, display_actual_end, output_file)
        print(f"HTML output generated: {html_file}")
        print(f"HTML shows lines {display_actual_start} to {display_actual_end}")
        
        # Check if p1.bat exists
        batch_file = 'p1.bat'
        if not os.path.exists(batch_file):
            print(f"Warning: Batch file '{batch_file}' not found")
            print("Output files created successfully, but batch file not executed")
            return
        
        # Run the batch file
        print(f"Running {batch_file}...")
        try:
            result = subprocess.run([batch_file], shell=True, capture_output=True, text=True)
            if result.returncode == 0:
                print("Batch file executed successfully")
                if result.stdout:
                    print("Output:", result.stdout)
            else:
                print(f"Batch file execution failed with return code: {result.returncode}")
                if result.stderr:
                    print("Error:", result.stderr)
        except Exception as e:
            print(f"Error running batch file: {str(e)}")
    
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main() 