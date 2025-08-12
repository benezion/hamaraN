#!/usr/bin/env python3
"""
Load Milon - Standalone Excel Processor
Processes milon Excel file and adds first words from multi-word phrases.
Maintains original formatting, colors, and structure.
"""

import pandas as pd
import os
import shutil
import warnings
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
import openpyxl.utils

# Suppress openpyxl warnings about unsupported image formats
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class MilonProcessor:
    def __init__(self, source_file, backup_dir="backup"):
        """
        Initialize the Milon processor
        
        Args:
            source_file (str): Path to the source Excel file
            backup_dir (str): Directory to store backups
        """
        self.source_file = source_file
        self.backup_dir = backup_dir
        self.working_file = None
        self.df = None
        
        # Create backup directory if it doesn't exist
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
            
        # Verify source file exists
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"Source file not found: {source_file}")
    
    def create_backup(self):
        """Create a timestamped backup of the original file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.basename(self.source_file)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(self.backup_dir, backup_filename)
        
        shutil.copy2(self.source_file, backup_path)
        print(f"✅ Backup created: {backup_path}")
        return backup_path
    
    def load_excel(self, sheet_name=None):
        """
        Load Excel file into DataFrame
        
        Args:
            sheet_name (str, optional): Sheet to load. If None, loads first sheet
        """
        try:
            # If no specific sheet is requested, load the first sheet (index 0)
            if sheet_name is None:
                self.df = pd.read_excel(self.source_file, sheet_name=0)
            else:
                self.df = pd.read_excel(self.source_file, sheet_name=sheet_name)
            
            print(f"✅ Loaded Excel file: {self.source_file}")
#            print(f"   Shape: {self.df.shape}")
            return True
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False
    
    def show_sample_data(self, n=5):
        """Display first n rows of the DataFrame"""
        if self.df is None:
            print("❌ No data loaded. Call load_excel() first.")
            return
        
#        print(f"\n--- Sample Data (first {n} rows) ---")
#        print(self.df.head(n))
#        print()
    
    def remove_empty_rows(self):
        """Remove completely empty rows"""
        if self.df is None:
            print("❌ No data loaded. Call load_excel() first.")
            return False
        
        original_count = len(self.df)
        self.df = self.df.dropna(how='all')
        new_count = len(self.df)
        
        print(f"✅ Removed empty rows: {original_count} → {new_count} rows ({original_count - new_count} removed)")
        return True
    
    def save_modified_file(self, output_file=None):
        """
        Save the modified DataFrame by copying original file and updating it
        This preserves all formatting, colors, and structure
        
        Args:
            output_file (str, optional): Output filename. If None, overwrites source file
        """
        if self.df is None:
            print("❌ No data loaded. Call load_excel() first.")
            return False

        if output_file is None:
            output_file = self.source_file

        try:
            # Step 1: Always copy original to preserve ALL formatting, colors, structure
            if output_file != self.source_file:
                shutil.copy2(self.source_file, output_file)
            
            # Step 2: Load the copied workbook
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Step 3: Smart data clearing - only clear actual data rows, preserve all formatting
            # Find actual data range (not empty rows at the end)
            actual_data_rows = len(self.df) + 1  # +1 for header
            max_data_col = len(self.df.columns) if not self.df.empty else 7
            
            # Only clear cells that will be overwritten, preserve all formatting
            # Clear from row 2 to the end of our new data + small buffer
            clear_until_row = min(ws.max_row, actual_data_rows + 50)  # Small buffer
            clear_until_col = min(ws.max_column, max_data_col + 3)    # Small buffer
            
            print(f"🧹 Clearing data area: rows 2-{clear_until_row}, cols 1-{clear_until_col}")
            
            # Surgically clear only the data area we'll use
            for row in range(2, clear_until_row + 1):  # Skip header row
                for col in range(1, clear_until_col + 1):
                    try:
                        cell = ws.cell(row=row, column=col)
                        cell.value = None  # Clear value but preserve formatting
                    except:
                        continue  # Skip problematic cells
            
            # Step 4: Update headers if we have new columns
            df_columns = list(self.df.columns)
            original_headers = [ws.cell(row=1, column=c).value for c in range(1, clear_until_col + 1)]
            
            # Add new column headers if needed
            for c_idx, col_name in enumerate(df_columns, start=1):
                ws.cell(row=1, column=c_idx).value = col_name
            
            # Step 5: Sort data by Column A for optimal binary search performance
            print("🔧 Optimizing data for fast binary search...")
            
            # Get column A name (should be the Hebrew source column)
            column_a_name = self.df.columns[0]  # Column A is the first column
            
            # Create sort key for Column A (same logic as translationLobe.py)
            self.df['_sort_key'] = (self.df[column_a_name]
                                   .astype(str)
                                   .str.strip()
                                   .str.replace("'", "׳")
                                   .str.lower())
            
            # Sort by the normalized key for Column A
            self.df = self.df.sort_values('_sort_key').reset_index(drop=True)
            
            print(f"✅ Data sorted by Column A ('{column_a_name}') for binary search optimization")
            
            # Step 6: Write the updated DataFrame data
            for r_idx, (_, row_data) in enumerate(self.df.iterrows(), start=2):
                for c_idx, (col_name, value) in enumerate(row_data.items(), start=1):
                    ws.cell(row=r_idx, column=c_idx).value = value
            
            # Step 7: Apply beautiful formatting before final save
            print("🎨 Applying beautiful formatting...")
            self._apply_nice_formatting(wb)
            
            # Step 8: Save the workbook with formatting
            wb.save(output_file)
            print("✅ Beautiful formatting applied successfully!")
            self.working_file = output_file
            return True
            
        except Exception as e:
            print(f"❌ Error saving file: {e}")
            return False
    
    def _apply_nice_formatting(self, wb):
        """Apply beautiful formatting to all worksheets"""
        try:
            # Define EXACT original color scheme from milon_zachar_nekeva.xlsx
            HEADER_COLOR_BLUE = "366092"     # Blue headers for Noun_Genders, Halacha, מספרים (FF366092)
            ALT_ROW_COLOR = "F2F2F2"         # Original light gray (FFF2F2F2) 
            WHITE_ROW_COLOR = "000000"       # Original white rows (00000000)
            BORDER_COLOR = "D0D0D0"          # Original border gray (FFD0D0D0)
            # Note: Main sheet (גיליון1) uses theme color, other sheets use blue RGB

            # Define styles - Updated per requirements
            header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")  # Headers: 12pt bold white
            data_font = Font(name="Arial", size=12)  # Data: 12pt (updated from 11pt)
            
            # Create different header fills for different sheets
            header_fill_blue = PatternFill(start_color=HEADER_COLOR_BLUE, end_color=HEADER_COLOR_BLUE, fill_type="solid")
            alt_row_fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
            white_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # True white for alternating

            thin_border = Border(
                left=Side(style='thin', color=BORDER_COLOR),
                right=Side(style='thin', color=BORDER_COLOR),
                top=Side(style='thin', color=BORDER_COLOR),
                bottom=Side(style='thin', color=BORDER_COLOR)
            )

            # Alignment - Updated per requirements
            center_alignment = Alignment(horizontal='center', vertical='center')  # For headers only
            right_alignment = Alignment(horizontal='right', vertical='center')    # For ALL data fields

            # Format each worksheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"  🎨 Formatting sheet: {sheet_name}")
                
                # Get actual data range with reasonable limits
                max_row = min(ws.max_row, 1000)  # Limit for performance
                max_col = min(ws.max_column, 15)
                
                # Format headers (row 1) - Different treatment per sheet
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    
                    # Apply different header colors based on sheet type
                    if sheet_name == 'גיליון1':
                        # Main sheet: Preserve original formatting (theme-based headers)
                        # Only set font and alignment, keep existing fill color
                        cell.font = Font(name="Arial", size=12, bold=True)  # No color override for main sheet
                    else:
                        # Other sheets: Use blue background with white text (like original)
                        cell.fill = header_fill_blue
                        cell.font = header_font  # White text for blue background
                
                # Format data rows - EXACT original pattern
                for row in range(2, max_row + 1):
                    # Apply row formatting with EXACT original alternating pattern
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        
                        # Set font (original uses Arial)
                        cell.font = data_font
                        
                        # EXACT original alternating: Even rows = gray (F2F2F2), Odd rows = white
                        if row % 2 == 0:  # Even rows (2,4,6,8...) = gray
                            cell.fill = alt_row_fill
                        else:  # Odd rows (3,5,7,9...) = white
                            cell.fill = white_row_fill
                        
                        # Set border and alignment
                        cell.border = thin_border
                        cell.alignment = right_alignment  # ALL data fields right-aligned
                
                # Set column widths
                for col in range(1, max_col + 1):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    if col == 1:  # Hebrew text column
                        ws.column_dimensions[column_letter].width = 20
                    elif col == 2:  # Nikud column
                        ws.column_dimensions[column_letter].width = 25
                    else:  # Other columns
                        ws.column_dimensions[column_letter].width = 15
                
                # Set row heights for first 100 rows
                for row in range(1, min(max_row + 1, 101)):
                    ws.row_dimensions[row].height = 20
            
            # Special formatting for main sheet
            if wb.sheetnames:
                ws_main = wb[wb.sheetnames[0]]  # First sheet (main data)
                # Add freeze panes and autofilter
                ws_main.freeze_panes = 'A2'
                if ws_main.max_row > 1:
                    ws_main.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws_main.max_column)}1"
                print(f"  ✅ Added freeze panes and autofilter to main sheet")
            
        except Exception as e:
            print(f"⚠️ Warning: Could not apply all formatting: {e}")
            # Continue without failing the entire process

def process_milon_file():
    """
    Main function to process the milon file and extract first words
    """
    print("🔧 Processing milon Excel file...")
    
    try:
        # Initialize the processor with your file
        processor = MilonProcessor("milon_zachar_nekeva.xlsx")
        
        # Create backup
        processor.create_backup()
        
        if not processor.load_excel():
            return False
        
        # Show file structure
        processor.show_sample_data(5)
        
        print("\n🔍 Analyzing multi-word phrases and extracting first words...")
        
        # Remove empty rows first
        processor.remove_empty_rows()
        
        # Get current DataFrame
        df = processor.df
        
        # Use first column for Hebrew words
        hebrew_column = df.columns[0]
        
        print(f"📊 Working with Hebrew column: '{hebrew_column}'")
        
        # Statistics tracking
        input_lines = len(df)
        lines_with_1_word = 0
        lines_with_2_words = 0
        lines_with_3_words = 0
        lines_with_more_words = 0
        inserted_from_2_words = 0
        inserted_from_3_words = 0
        inserted_from_more_words = 0
        
        # Collect existing single words (to avoid duplicates)
        existing_words = set()
        for _, row in df.iterrows():
            hebrew_text = str(row[hebrew_column]).strip()
            if hebrew_text and hebrew_text != 'nan':
                words = hebrew_text.split()
                if len(words) == 1:
                    existing_words.add(words[0])
                    lines_with_1_word += 1
                elif len(words) == 2:
                    lines_with_2_words += 1
                elif len(words) == 3:
                    lines_with_3_words += 1
                else:
                    lines_with_more_words += 1
        
        print(f"📈 Found {len(existing_words)} existing single words")
        
        # Now extract first words from multi-word phrases
        new_rows = []
        
        for _, row in df.iterrows():
            hebrew_text = str(row[hebrew_column]).strip()
            if hebrew_text and hebrew_text != 'nan':
                words = hebrew_text.split()
                word_count = len(words)
                
                # Only process multi-word phrases
                if word_count >= 2:
                    first_word = words[0]
                    
                    # Check if first word already exists as single entry
                    if first_word not in existing_words:
                        # Create new row with only Column A and Column G (Special) filled
                        # All other columns will be empty as requested
                        new_row = pd.Series(index=df.columns, dtype=object)
                        # Fill all columns with empty strings
                        new_row = new_row.fillna('')
                        
                        # Set only Column A (Hebrew source) and Special column
                        new_row[hebrew_column] = first_word  # Column A: First word
                        new_row['Special'] = 'Special'       # Column G: Special marker
                        
                        new_rows.append(new_row)
                        
                        # Add to existing words to avoid future duplicates
                        existing_words.add(first_word)
                        
                        # Track insertion statistics
                        if word_count == 2:
                            inserted_from_2_words += 1
                        elif word_count == 3:
                            inserted_from_3_words += 1
                        else:
                            inserted_from_more_words += 1
                        
 #                       print(f"➕ Adding: '{first_word}' from '{hebrew_text}'")
        
        # Add new rows to the DataFrame
        if new_rows:
            # Get original column order
            original_cols = list(processor.df.columns)
            
            # Add new rows
            new_df = pd.DataFrame(new_rows)
            processor.df = pd.concat([processor.df, new_df], ignore_index=True)
            
            # Now reorder columns to put Special at the very end, removing any gaps
            all_cols = list(processor.df.columns)
            
            # Remove Special from wherever it is
            if 'Special' in all_cols:
                all_cols.remove('Special')
            
            # Remove any empty or unnamed columns that might cause gaps
            cleaned_cols = []
            for col in all_cols:
                col_str = str(col).strip()
                if col_str and col_str != 'nan' and not col_str.startswith('Unnamed'):
                    cleaned_cols.append(col)
            
            # Add Special at the end
            final_col_order = cleaned_cols + ['Special']
            
            # Reorder the DataFrame
            processor.df = processor.df[final_col_order]
            
            # Fill NaN values in Special column
            processor.df['Special'] = processor.df['Special'].fillna('')
            
            print(f"✅ Added {len(new_rows)} new first-word entries")
        else:
            print("ℹ️ No new first-word entries needed")
        
        # Statistics
        total_inserted = inserted_from_2_words + inserted_from_3_words + inserted_from_more_words
        total_export_rows = input_lines + total_inserted
        
        print("\n" + "="*60)
        print("📊 PROCESSING STATISTICS:")
        print("="*60)
        print(f"📥 Input lines:                    {input_lines:,}")
        print(f"📝 Lines with 1 word:              {lines_with_1_word:,}")
        print(f"📝 Lines with 2 words:             {lines_with_2_words:,}")
        print(f"📝 Lines with 3 words:             {lines_with_3_words:,}")
        print(f"📝 Lines with 4+ words:            {lines_with_more_words:,}")
        print("-" * 60)
        print(f"➕ Inserted from 2-word phrases:   {inserted_from_2_words:,}")
        print(f"➕ Inserted from 3-word phrases:   {inserted_from_3_words:,}")
        print(f"➕ Inserted from 4+-word phrases:  {inserted_from_more_words:,}")
        print(f"➕ Total inserted lines:           {total_inserted:,}")
        print("-" * 60)
        print(f"📤 Total export rows:              {total_export_rows:,}")
        print("="*60)
        
        
        # Save the modified file (maintains original formatting)
        output_filename = "milon_zachar_nekeva_new.xlsx"
        if processor.save_modified_file(output_filename):
            return True
        else:
            return False
            
    except Exception as e:
        return False

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "--help":
        print("Load Milon - Excel Import/Export Processor")
        print("Usage:")
        print("  python load_milon.py        # Process milon file and create modified version")
        print("  python load_milon.py --help # Show this help")
        print()
        print("Input:  milon_zachar_nekeva.xlsx")
        print("Output: milon_zachar_nekeva_new.xlsx")
        print()
        print("The program:")
        print("- Analyzes multi-word Hebrew phrases")
        print("- Extracts first words that don't exist as single entries")
        print("- Adds them with 'Special' marker")
        print("- Maintains original Excel formatting and structure")
    else:
        # Run the processing
        success = process_milon_file()